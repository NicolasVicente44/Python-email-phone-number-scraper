# This program grabs all phone numbers and emails from the users clipboard allowing them to quickly obtain them. Additionally the program writes the grabbed info and saves it to an excel workbook.

import pyperclip # pyperclip is a clipboard module for python that allows you to manipulate clipboard data
import re # import statement for regex
from openpyxl import Workbook # imports the excel functionality to the program


# Get clipboard content
clipboard = pyperclip.paste()





#the verbose method allows you to add comments and whitespace to the regex to make it more readable and modifiable
# the regular expression pattern for phone numbers
phoneRegex = re.compile(r'''(
    (\d{3}|\(\d{3}\))? # area code
    (\s|-|\.)? # separator
    (\d{3}) # first 3 digits
    (\s|-|\.) # separator
    (\d{4}) # last 4 digits
    (\s*(ext|x|ext.)\s*(\d{2,5}))? # extension
    )''', re.VERBOSE)

#the regular expression pattern for emails
emailRegex = re.compile(r'''(
    [a-zA-Z0-9._%+-]+ # username
    @ # @ symbol
    [a-zA-Z0-9.-]+ # domain name
    (\.[a-zA-Z]{2,4}) # dot-something
    )''', re.VERBOSE)





areaCodes = [] #set of unqiue area codes


# Search for phone numbers and email addresses, if they exisst add them to each respective list
phoneNumbers = []
emailAddresses = []


# append the data that matches the regexs in all the clipboard content to the respective lists
for match in phoneRegex.findall(clipboard):
    phoneNumbers.append(match[0])
    for phoneNumber in phoneNumbers:
        areaCode = phoneNumber[:3] # extract the first 3 digits of the phone number
        if areaCode not in areaCodes: #if the area code is unique, append it to the list
            areaCodes.append(areaCode)
print("Unique Area Codes:")
for i in range(len(areaCodes)): #loops over the area codes list and print them
    print(areaCodes[i])

    
    
for match in emailRegex.findall(clipboard):
    emailAddresses.append(match[0])









wb = Workbook()  #creates a new excel workbook

ws = wb.active # get the active excel work sheet







# if a phone number or email was grabbed, write them to an excel sheet
if phoneNumbers:
    ws['A1'] = 'Phone Numbers' # title the A1 cell/column phone numbers
    for i, phoneNumber in enumerate(phoneNumbers):
        ws.cell(row = i + 2, column = 1).value = phoneNumber
        
        
if emailAddresses:
    ws['B1'] = 'Email Addresses' # title the B1 cell/column email addresses
    for i, emailAddress in enumerate(emailAddresses):
        ws.cell(row = i + 2, column = 2).value = emailAddress






wb.save('grabbed_data.xlsx')  # saves the worksheet in the same directory/folder as my python code








# copy extracted data to clipboard
output = "Phone Numbers:\n" + "\n".join(phoneNumbers) + "\n\nEmail Addresses:\n" + "\n".join(emailAddresses)
pyperclip.copy(output)






#this function gives the user feedback 
def feedBack ():

    #give feedback based on if phone numbers or emails are found in the selection, if none are found provide a helpful message
    if len(phoneNumbers) == 0:
        print("No phone number found in that selection")
    else:
        print(f"{len(phoneNumbers)} phone numbers found in that selection")


    if len(emailAddresses) == 0:
        print("No email addresses found in that selection")
    else:
        print(f"{len(emailAddresses)} email addresses found in that selection")
        
        
feedBack()